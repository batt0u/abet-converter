from pathlib import Path

import pytest
from openpyxl import load_workbook

import abet_converter.components.ingest.abet_to_sqlite as ingest_module
from abet_converter.components.ingest.abet_to_sqlite import (
    ConversionTarget,
    build_conversion_targets,
    convert_one,
    discover_abet_sources,
    ensure_target_paths_do_not_exist,
    normalize_formats,
    normalize_sheet_name,
    parse_csv_text,
    write_table_csv,
    write_xlsx_workbook,
)
from abet_converter.drivers.mdbtools import MdbtoolsRuntime


def test_discover_abet_sources_reads_file_and_folder(tmp_path: Path) -> None:
    abet = tmp_path / "a.ABETdb"
    mdb = tmp_path / "b.mdb"
    other = tmp_path / "notes.txt"
    abet.write_text("x", encoding="utf-8")
    mdb.write_text("x", encoding="utf-8")
    other.write_text("x", encoding="utf-8")

    assert discover_abet_sources(abet) == [abet.resolve()]
    assert discover_abet_sources(tmp_path) == [abet.resolve(), mdb.resolve()]


def test_discover_abet_sources_supports_recursive_search(tmp_path: Path) -> None:
    top_level = tmp_path / "top.ABETdb"
    nested_dir = tmp_path / "nested"
    nested_dir.mkdir()
    nested = nested_dir / "nested.mdb"
    top_level.write_text("x", encoding="utf-8")
    nested.write_text("x", encoding="utf-8")

    assert discover_abet_sources(tmp_path, recursive=False) == [top_level.resolve()]
    assert discover_abet_sources(tmp_path, recursive=True) == [nested.resolve(), top_level.resolve()]


def test_discover_abet_sources_rejects_unsupported_file(tmp_path: Path) -> None:
    source = tmp_path / "notes.txt"
    source.write_text("x", encoding="utf-8")

    with pytest.raises(ValueError, match="Unsupported source extension"):
        discover_abet_sources(source)


def test_normalize_formats_deduplicates_and_preserves_order() -> None:
    assert normalize_formats(["CSV", "sqlite", "csv", "XLSX"]) == ("csv", "sqlite", "xlsx")


def test_normalize_formats_rejects_unknown_format() -> None:
    with pytest.raises(ValueError, match="Unsupported format"):
        normalize_formats(["json"])


def test_build_conversion_targets_for_single_file_sqlite_only_uses_explicit_file(tmp_path: Path) -> None:
    source = tmp_path / "input.ABETdb"
    source.write_text("x", encoding="utf-8")

    targets = build_conversion_targets(
        sources=[source.resolve()],
        input_path=source.resolve(),
        output_path=(tmp_path / "output.sqlite").resolve(),
        formats=("sqlite",),
    )

    assert targets == [
        ConversionTarget(
            source_db=source.resolve(),
            sqlite_path=(tmp_path / "output.sqlite").resolve(),
            sql_dump_path=None,
            csv_dir=None,
            xlsx_path=None,
        )
    ]


def test_build_conversion_targets_for_single_file_multi_format_uses_output_directory(tmp_path: Path) -> None:
    source = tmp_path / "input.ABETdb"
    source.write_text("x", encoding="utf-8")

    targets = build_conversion_targets(
        sources=[source.resolve()],
        input_path=source.resolve(),
        output_path=(tmp_path / "exports").resolve(),
        formats=("sqlite", "sql", "csv", "xlsx"),
    )

    assert targets == [
        ConversionTarget(
            source_db=source.resolve(),
            sqlite_path=(tmp_path / "exports" / "input.sqlite").resolve(),
            sql_dump_path=(tmp_path / "exports" / "input.sql").resolve(),
            csv_dir=(tmp_path / "exports" / "input_csv").resolve(),
            xlsx_path=(tmp_path / "exports" / "input.xlsx").resolve(),
        )
    ]


def test_build_conversion_targets_for_directory_preserves_relative_structure(tmp_path: Path) -> None:
    input_dir = tmp_path / "incoming"
    nested_dir = input_dir / "nested"
    nested_dir.mkdir(parents=True)
    source = nested_dir / "sample.ABETdb"
    source.write_text("x", encoding="utf-8")

    targets = build_conversion_targets(
        sources=[source.resolve()],
        input_path=input_dir.resolve(),
        output_path=(tmp_path / "converted").resolve(),
        formats=("sqlite", "sql", "csv", "xlsx"),
    )

    assert targets == [
        ConversionTarget(
            source_db=source.resolve(),
            sqlite_path=(tmp_path / "converted" / "nested" / "sample.sqlite").resolve(),
            sql_dump_path=(tmp_path / "converted" / "nested" / "sample.sql").resolve(),
            csv_dir=(tmp_path / "converted" / "nested" / "sample_csv").resolve(),
            xlsx_path=(tmp_path / "converted" / "nested" / "sample.xlsx").resolve(),
        )
    ]


def test_parse_csv_text_handles_empty_input() -> None:
    assert parse_csv_text("") == ([], [])


def test_parse_csv_text_converts_null_sentinel_and_keeps_empty_string() -> None:
    headers, rows = parse_csv_text('ID,Name,Note\n1,Alpha,__NULL__\n2,,empty-name\n')

    assert headers == ["ID", "Name", "Note"]
    assert rows == [["1", "Alpha", None], ["2", "", "empty-name"]]


def test_write_table_csv_writes_none_as_empty_cell(tmp_path: Path) -> None:
    csv_path = tmp_path / "nested" / "table.csv"

    write_table_csv(csv_path, ["ID", "Note"], [["1", None], ["2", "ok"]])

    assert csv_path.read_text(encoding="utf-8").splitlines() == ["ID,Note", "1,", "2,ok"]


def test_normalize_sheet_name_truncates_and_deduplicates() -> None:
    used_names: set[str] = set()

    first = normalize_sheet_name("x" * 40, used_names)
    second = normalize_sheet_name("x" * 40, used_names)
    blank = normalize_sheet_name("", used_names)

    assert first == "x" * 31
    assert second == ("x" * 29) + "_1"
    assert blank == "Sheet"


def test_write_xlsx_workbook_splits_large_tables_across_sheets(tmp_path: Path) -> None:
    xlsx_path = tmp_path / "out.xlsx"

    write_xlsx_workbook(
        xlsx_path,
        [("tbl_Data", ["ID"], [["1"], ["2"], ["3"], ["4"], ["5"]])],
        max_rows_per_sheet=3,
    )

    workbook = load_workbook(xlsx_path, read_only=True)
    try:
        assert workbook.sheetnames == ["tbl_Data", "tbl_Data_2", "tbl_Data_3"]
        assert [row for row in workbook["tbl_Data"].iter_rows(values_only=True)] == [("ID",), ("1",), ("2",)]
        assert [row for row in workbook["tbl_Data_2"].iter_rows(values_only=True)] == [("ID",), ("3",), ("4",)]
        assert [row for row in workbook["tbl_Data_3"].iter_rows(values_only=True)] == [("ID",), ("5",)]
    finally:
        workbook.close()


def test_ensure_target_paths_do_not_exist_rejects_existing_csv_dir(tmp_path: Path) -> None:
    csv_dir = tmp_path / "out_csv"
    csv_dir.mkdir()
    target = ConversionTarget(
        source_db=tmp_path / "input.ABETdb",
        sqlite_path=None,
        sql_dump_path=None,
        csv_dir=csv_dir,
        xlsx_path=None,
    )

    with pytest.raises(FileExistsError, match="Output already exists"):
        ensure_target_paths_do_not_exist(target)


def test_convert_one_refuses_existing_sqlite(tmp_path: Path) -> None:
    sqlite_path = tmp_path / "out.sqlite"
    sqlite_path.write_text("already here", encoding="utf-8")
    target = ConversionTarget(
        source_db=tmp_path / "input.ABETdb",
        sqlite_path=sqlite_path,
        sql_dump_path=None,
        csv_dir=None,
        xlsx_path=None,
    )
    runtime = MdbtoolsRuntime(bundle_dir=tmp_path / "runtime", executable_suffix=".exe", library_env_var=None)

    with pytest.raises(FileExistsError):
        convert_one(target, runtime)


def test_convert_one_writes_requested_formats(monkeypatch: pytest.MonkeyPatch, tmp_path: Path) -> None:
    source = tmp_path / "input.ABETdb"
    source.write_text("x", encoding="utf-8")
    runtime_dir = tmp_path / "runtime"
    runtime_dir.mkdir()
    for executable in ("mdb-tables.exe", "mdb-schema.exe", "mdb-export.exe"):
        (runtime_dir / executable).write_text("stub", encoding="utf-8")
    runtime = MdbtoolsRuntime(bundle_dir=runtime_dir, executable_suffix=".exe", library_env_var=None)

    monkeypatch.setattr(ingest_module, "get_tables", lambda *_args: ["tbl_Schedules"])
    monkeypatch.setattr(
        ingest_module,
        "get_schema_sql",
        lambda *_args: 'CREATE TABLE "tbl_Schedules" ("SID" TEXT, "Name" TEXT);',
    )
    monkeypatch.setattr(ingest_module, "export_table_csv", lambda *_args: "SID,Name\n1,Alpha\n2,__NULL__\n")

    target = ConversionTarget(
        source_db=source,
        sqlite_path=tmp_path / "out.sqlite",
        sql_dump_path=tmp_path / "out.sql",
        csv_dir=tmp_path / "out_csv",
        xlsx_path=tmp_path / "out.xlsx",
    )

    written_paths = convert_one(target, runtime)

    assert target.sqlite_path in written_paths
    assert target.sql_dump_path in written_paths
    assert target.xlsx_path in written_paths
    assert (target.csv_dir / "tbl_Schedules.csv") in written_paths
    assert target.sqlite_path.exists()
    assert target.sql_dump_path.exists()
    assert target.xlsx_path.exists()
    assert (target.csv_dir / "tbl_Schedules.csv").exists()
    assert "INSERT INTO" in target.sql_dump_path.read_text(encoding="utf-8")


def test_convert_one_raises_when_source_has_no_tables(monkeypatch: pytest.MonkeyPatch, tmp_path: Path) -> None:
    source = tmp_path / "input.ABETdb"
    source.write_text("x", encoding="utf-8")
    runtime = MdbtoolsRuntime(bundle_dir=tmp_path / "runtime", executable_suffix=".exe", library_env_var=None)
    target = ConversionTarget(
        source_db=source,
        sqlite_path=tmp_path / "out.sqlite",
        sql_dump_path=None,
        csv_dir=None,
        xlsx_path=None,
    )
    monkeypatch.setattr(ingest_module, "get_tables", lambda *_args: [])

    with pytest.raises(RuntimeError, match="No tables found"):
        convert_one(target, runtime)


def test_convert_one_does_not_create_sqlite_when_not_requested(monkeypatch: pytest.MonkeyPatch, tmp_path: Path) -> None:
    source = tmp_path / "input.ABETdb"
    source.write_text("x", encoding="utf-8")
    runtime_dir = tmp_path / "runtime"
    runtime_dir.mkdir()
    for executable in ("mdb-tables.exe", "mdb-schema.exe", "mdb-export.exe"):
        (runtime_dir / executable).write_text("stub", encoding="utf-8")
    runtime = MdbtoolsRuntime(bundle_dir=runtime_dir, executable_suffix=".exe", library_env_var=None)

    monkeypatch.setattr(ingest_module, "get_tables", lambda *_args: ["tbl_Schedules"])
    monkeypatch.setattr(
        ingest_module,
        "get_schema_sql",
        lambda *_args: 'CREATE TABLE "tbl_Schedules" ("SID" TEXT);',
    )
    monkeypatch.setattr(ingest_module, "export_table_csv", lambda *_args: "SID\n1\n")

    target = ConversionTarget(
        source_db=source,
        sqlite_path=None,
        sql_dump_path=tmp_path / "out.sql",
        csv_dir=tmp_path / "out_csv",
        xlsx_path=tmp_path / "out.xlsx",
    )

    convert_one(target, runtime)

    assert not (tmp_path / "input.sqlite").exists()
    assert target.sql_dump_path.exists()
    assert target.xlsx_path.exists()
    assert (target.csv_dir / "tbl_Schedules.csv").exists()
